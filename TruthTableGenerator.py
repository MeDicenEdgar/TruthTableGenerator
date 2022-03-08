
from importlib.resources import path
from tkinter import Frame, Tk, Label, Text, Button, END
from tkinter.constants import BOTH
import openpyxl
from openpyxl import Workbook
from openpyxl.cell import cell
from openpyxl.workbook.workbook import Workbook
from tkinter import filedialog
import re
#Importameos TKinter para la interfaz, openpyxl para la descarga y re para el encuentro de parentesis

window = Tk()
window.title("Genedador de tablas de verdad")
window.geometry("329x400")
inputFrame = Frame(window, width=400, height=600, background="#210042")
resultsFrame = Frame(window, width=400, height=600, background="#210042")
screen = Text(window, state="disabled", width=17, height=2, background="white", foreground="black", font=("Arial",25))
declaration = ""
screen.place(x=10,y=15)
resultsWB = Workbook() 
resultspg = resultsWB.active 
#variables obligatorias de tkinter y de openpyxl

truthTable = [[]]
numbers = set()
compare = False
compareList = []
#Generamos la tabla de verdad y la variable de comparacion para poder saber si ya se ejecutó el codigo una vez
#Asi como generamo un set de numeros para saber que indices se puede usar y evitar problemas con el eval

def nope(index):#la función del no, niega todos los indices y los añade a la tabla de verdad
    global truthTable
    newIndex = []
    for i in range(len(truthTable[1])):
        newIndex.append(not(truthTable[index][i]))
    truthTable.append(newIndex)
    return(len(truthTable)-1)#te regresa el indice donde lo guardó

def andu(index1, index2):#lo mismo que el no pero te con 2 indices, haciendo una comparacion logica entre los 2
    global truthTable
    newIndex = []
    for i in range(len(truthTable[1])):
        newIndex.append(truthTable[index1][i] and truthTable[index2][i])
    truthTable.append(newIndex)
    return(len(truthTable)-1)

def oru(index1, index2):#lo mismo que el no pero te con 2 indices, haciendo una comparacion logica entre los 2
    global truthTable
    newIndex = []
    for i in range(len(truthTable[1])):
        newIndex.append(truthTable[index1][i] or truthTable[index2][i])
    truthTable.append(newIndex)
    return(len(truthTable)-1)

def then(index1, index2):#lo mismo que el no pero te con 2 indices, haciendo una comparacion por el medio de ifs
    global truthTable
    newIndex = []
    for i in range(len(truthTable[1])):
        if truthTable[index1][i] == False:
            newIndex.append(True)
        else:
            if truthTable[index2][i] == True:
                newIndex.append(True)
            else:
                newIndex.append(False)
    truthTable.append(newIndex)
    return(len(truthTable)-1)

def onlyIf(index1, index2):#lo mismo que el no pero te con 2 indices, haciendo una comparacion por el medio de ifs
    global truthTable
    newIndex = []
    for i in range(len(truthTable[1])):
        if truthTable[index1][i] == True and truthTable[index2][i]== True:
            newIndex.append(True)
        elif truthTable[index1][i] == False and truthTable[index2][i]== False:
            newIndex.append(True)
        else:
            newIndex.append(False)
    truthTable.append(newIndex)
    return(len(truthTable)-1)

def createButton(valor, ancho=5, alto=2):#crea los botones y almacena el valor en write
        return Button(inputFrame, text=valor, width=ancho, height=alto, font=("Helvetica",15), command=lambda:write(valor))
    
def write(valor):#Utiliza el valor para modificar la preposición y escribir en la pantalla
    global declaration
    declaration += valor
    screen.configure(state="normal")
    screen.insert(END, valor)
    screen.configure(state="disabled")

def delete():#resetea la preposición y la pantalla
    global declaration
    declaration = ""
    screen.configure(state="normal")
    screen.delete("1.0", END)
    screen.configure(state="disabled")

def start(declaration, compare, compareList):#La funcion que hace todo el proceso
    global truthTable
    inputs = 0
    
    if "p" in declaration:
        truthTable[0].append("p")
        inputs += 1 
        numbers.add("1")
        p = 1
    if "q" in declaration:
        if inputs == 0:
            q = 1
            numbers.add("1")
        else:
            q = 2
            numbers.add("2")
        truthTable[0].append("q")
        inputs += 1
    if "r" in declaration:#Determina cuantos inputs tenemos y con eso genera las columnas default, las variables p q y r son los indices de cual columnac orresponde a cada uno
        if inputs == 0:
            r = 1
            numbers.add("1")
        elif inputs == 1:
            r = 2
            numbers.add("2")
        elif inputs==2:
            r = 3
            numbers.add("3")
        truthTable[0].append("r")
        inputs += 1
    #Generamos los inputs con variables que nos dicen en que columna de la tabla de verdad se encuentra cada una, asi como añadirlos a la
    #lista de nombres de nuestra tabla
    if inputs == 0:
        print("That is not a valid expression")
    elif inputs == 1:
        truthTable.append([True, False])
    elif inputs == 2:
        truthTable.append([True, True, False, False])
        truthTable.append([True, False, True, False])
    elif inputs == 3:
        truthTable.append([True, True, False, False, True, True, False, False])
        truthTable.append([True, False, True, False, True, False, True, False])
        truthTable.append([True, True, True, True, False, False, False, False])
    #Determina como se representan las funciones primitivas dependiendo de cuantas hay

    prepositions = re.findall('\(.*?\)', declaration)
    prepositions.insert(0, declaration)
    indexes = []#Nos ayuda a mantener un registro de que preposición corresponde a cual indice
    values = []
    #generamos un stack con las proposiciones en orden de prioridad y añadimos la proposición original como la ultima a resolver

    for element in reversed(prepositions):#Corremos los elementos de nuestro stack en reveersed, para que sea un stack
        element = element.replace("(", "")
        element = element.replace(")", "")
        indexes.append(element)#Quitamos los parentesis y cada que empiece la lista lo añadimos a los indices
        for i in range(len(indexes)-1):
            if indexes[i] in element:
                element=element.replace(indexes[i], values[i])#Analiza si hay alguna proposicion pasada en la nueva proposicion 
                                                              #y la reemplaza por su indice correspondiente

        while True:#Este true no se va a terminar hasta que la proposicion haya sido terminada de resolver
            if "~" in element:#Checa si hay negacion
                temp = element.find("~") 
                temp1 = temp+1#Encuentra donde está y el elemento que la acompaña
                if (len(element[temp:]))>2:
                    if element[temp+2] in numbers:
                        temp1 = element[temp+1]+element[temp+2]#Estos bloques de aqui son solo por si se da la casualidad de que el indice es
                if element[temp1] in numbers:                  #mayor que 10
                    index = nope(eval(element[temp1]))
                else:
                    index = nope(locals()[element[temp1]])#Revisa si la proposicion que está operando es primitiva o un indice de una compuesta
                for i in reversed(prepositions): #Al acabar y resolver reemplaza todas las apareciones de dicha proposición y las reemplaza por 
                    i = i.replace((element[temp]+element[temp1]), str(index))#el indice en el cual está almacenado
                newPreposition = element[temp]+element[temp1]
                for i in newPreposition:
                    if i in numbers:
                        newPreposition = newPreposition.replace(i, truthTable[0][eval(i)-1])
                truthTable[0].append(newPreposition)#Añade la proposición original a la primera lista de nuestra tabla, la cual es la de los nombres
                element = element.replace((element[temp]+element[temp1]), str(index))
                numbers.add(str(index))#Añade el indice a la lista de numeros
                continue
            
            if "∧" in element:#Basicamente la estructura de todos los checks es la misma pero cambiando la funcion
                temp = element.find("∧")
                temp1 = element[temp-1]
                temp2 = element[temp+1]

                if len(element[:temp]) > 2:
                    if element[temp-2] in numbers:
                        temp1 = element[temp-2]+element[temp-1]

                if len(element[temp:])>2:
                    print(len(element[temp:]))
                    if element[temp+2] in numbers:
                        temp2 = element[temp+1]+element[temp+2]#Lo mismo, checamos que no sea un mayor a 10 el indice

                if temp1 in numbers:
                    if temp2 in numbers:
                        index = andu(eval(element[temp-1]), eval(element[temp+1]))
                    else:
                        index = andu(eval(element[temp-1]), locals()[element[temp+1]])
                else:
                    if temp2 in numbers:
                        index = andu(locals()[element[temp-1]], eval(element[temp+1]))
                    else:
                        index = andu(locals()[element[temp-1]], locals()[element[temp+1]])#checamos si nos pasó indices o primitivos o una y una
                newPreposition = element[temp-1:temp+2]
                for i in newPreposition:
                    if i in numbers:
                        newPreposition = newPreposition.replace(i, truthTable[0][eval(i)-1])
                truthTable[0].append(newPreposition)#Reemplazamos para la lista de nombres
                element = element.replace((element[temp-1:temp+2]), str(index))
                numbers.add(str(index))#Añadimos el indice
                continue

            if "v" in element:#Lo mismo que el y pero llamando la funcion de or
                temp = element.find("v")
                temp1 = element[temp-1]
                temp2 = element[temp+1]
                if len(element[:temp]) > 2:
                    if element[temp-2] in numbers:
                        temp1 = element[temp-2]+element[temp-1]

                if len(element[temp:])>2:
                    print(len(element[temp:]))
                    if element[temp+2] in numbers:
                        temp2 = element[temp+1]+element[temp+2]
                if temp1 in numbers:
                    if temp2 in numbers:
                        index = oru(eval(element[temp-1]), eval(element[temp+1]))
                    else:
                        index = oru(eval(element[temp-1]), locals()[element[temp+1]])
                else:
                    if temp2 in numbers:
                        index = oru(locals()[element[temp-1]], eval(element[temp+1]))
                    else:
                        index = oru(locals()[element[temp-1]], locals()[element[temp+1]])
                newPreposition = element[temp-1:temp+2]
                for i in newPreposition:
                    if i in numbers:
                        newPreposition = newPreposition.replace(i, truthTable[0][eval(i)-1])
                truthTable[0].append(newPreposition)
                element = element.replace((element[temp-1:temp+2]), str(index))
                numbers.add(str(index))
                continue

            if "→" in element: #Lo mismo pero llamando la funcion del then
                temp = element.find("→")
                temp1 = element[temp-1]
                temp2 = element[temp+1]
                print(element[:temp])
                if len(element[:temp])+1 > 2:
                    if element[temp-2] in numbers:
                        temp1 = element[temp-2]+element[temp-1]

                if len(element[temp:])>2:
                    print(len(element[temp:]))
                    if element[temp+2] in numbers:
                        temp2 = element[temp+1]+element[temp+2]

                if temp1 in numbers:
                    if temp2 in numbers:
                        index = then(eval(element[temp-1]), eval(element[temp+1]))
                    else:
                        index = then(eval(element[temp-1]), locals()[element[temp+1]])
                else:
                    if temp2 in numbers:
                        index = then(locals()[element[temp-1]], eval(element[temp+1]))
                    else:
                        index = then(locals()[element[temp-1]], locals()[element[temp+1]])
                newPreposition = element[temp-1:temp+2]
                for i in newPreposition:
                    if i in numbers:
                        newPreposition = newPreposition.replace(i, truthTable[0][eval(i)-1])
                truthTable[0].append(newPreposition)
                element = element.replace((element[temp-1:temp+2]), str(index))
                numbers.add(str(index))
                continue

            if "↔" in element:#Lo mismo pero llamando el onlyif
                temp = element.find("↔")
                temp1 = element[temp-1]
                temp2 = element[temp+1]
                if len(element[:temp]) > 2:
                    if element[temp-2] in numbers:
                        temp1 = element[temp-2]+element[temp-1]

                if len(element[temp:])>2:
                    print(len(element[temp:]))
                    if element[temp+2] in numbers:
                        temp2 = element[temp+1]+element[temp+2]

                if temp1 in numbers:
                    if temp2 in numbers:
                        index = onlyIf(eval(element[temp-1]), eval(element[temp+1]))
                    else:
                        index = onlyIf(eval(element[temp-1]), locals()[element[temp+1]])
                else:
                    if temp2 in numbers:
                        index = onlyIf(locals()[element[temp-1]], eval(element[temp+1]))
                    else:
                        index = onlyIf(locals()[element[temp-1]], locals()[element[temp+1]])
                newPreposition = element[temp-1:temp+2]
                for i in newPreposition:
                    if i in numbers:
                        newPreposition = newPreposition.replace(i, truthTable[0][eval(i)-1])
                truthTable[0].append(newPreposition)
                element = element.replace((element[temp-1:temp+2]), str(index))
                numbers.add(str(index))
                continue
            values.append(element)     
            break
    inputFrame.pack_forget()
    resultsFrame.pack(fill=BOTH, expand=True)
    button13.place(x=22, y =180)
    deter = 0#cambiamos a la siguienta pantalla
    if compare == True:#Si este proceso es el segundo que hacemos y es para una comparacion, entra a este if
        equal = True
        if len(compareList) == len(truthTable[len(truthTable)-1]):#si no son de la misma longitus, por default se descarta la comparacion
            for i in range(len(compareList)-1):
                if compareList[i] == truthTable[len(truthTable)-1][i]:#checa si todos son iguales, sino se pone falso y te dice que no
                    pass
                else: 
                    equal = False
            if equal == True:
                screen.configure(state="normal")
                screen.insert(END, "\nSon iguales")
                screen.configure(state="disabled")
            else:
                screen.configure(state="normal")
                screen.insert(END, "\nNo son iguales")
                screen.configure(state="disabled")
        else:
            screen.configure(state="normal")
            screen.insert(END, "\nNo son iguales")#variables de tkinter para que se muestre que no son iguales
            screen.configure(state="disabled")
        button14.pack_forget()
    else:
        for n in truthTable[len(truthTable)-1]:
            if n == True:
                deter +=1
            if n == False:
                deter -=1
        if deter == len(truthTable[len(truthTable)-1]):
            screen.configure(state="normal")
            screen.insert(END, "\nTautología")#Si no es una comparacion, checa lo de taulogogía y contradiccion y lo muestra
            screen.configure(state="disabled")
        elif deter == ((len(truthTable[len(truthTable)-1]))*-1):
            screen.configure(state="normal")
            screen.insert(END, "\nContradicción")
            screen.configure(state="disabled")
        button14.place(x=22, y =280)

def download():#Esta funcion almacena todo en un excel y te lo da a descargar
    global truthTable
    
    for i in range(len(truthTable[0])):
        newCell = resultspg.cell(row=1, column=i+1)
        newCell.value = truthTable[0][i]#Imprime primero la lista de nombres
    for i in range(len(truthTable)):
        if i == 0:
            continue
        resultspg.column_dimensions[chr(64+i)].width = 13
        for n in range(len(truthTable[i])):
            newCell = resultspg.cell(row = n+2, column = i)
            newCell.value = truthTable[i][n]#Imprime abajo de la lista de nombres por columnas, la lista que le corresponde a cada uno
    path="{}/resultados.xlsx".format(filedialog.askdirectory(initialdir="/", title="Seleccione donde desea el archivo de resultados"))
    resultsWB.save(path)#te pregunta donde lo quieres guardar y te lo guarda

def contrast():#esta funcion resetea las variables globales, guarda el ultimo indice de la tabla y te manda a la pantalla inicial
    global compare
    global compareList
    global numbers
    global truthTable
    compareList = truthTable[len(truthTable)-1]
    truthTable=[[]]
    numbers = set()
    compare = True
    resultsFrame.pack_forget()
    inputFrame.pack()
    delete()#llama la funcion delete para resetear la pantalla y la preposición



 #De aqui para abajo solo es interfaz      

button1= createButton("p")
button2= createButton("q")
button3= createButton("r")
button4= createButton("~")
button5= createButton("∧")
button6= createButton("v")
button7= createButton("→")
button8= createButton("↔")
button9= createButton("(")
button10= createButton(")")
button11= Button(inputFrame, text="=", width=5, height=2, font=("Helvetica",15), command=lambda:start(declaration, compare, compareList))
button12= Button(inputFrame, text="AC", width=5, height=2, font=("Helvetica",15), command=lambda:delete())
button13= Button(resultsFrame, text="Descargar tabla de verdad", width=25, height=2, font=("Helvetica",15), command=lambda:download())
button14= Button(resultsFrame, text="Comparar con otra funcion", width=25, height=2, font=("Helvetica",15), command=lambda:contrast())



button1.place(x=10, y=120)
button2.place(x= 90, y =120)
button3.place(x=170, y=120)
button4.place(x= 250, y =120)
button5.place(x=10, y=205)
button6.place(x= 90, y =205)
button7.place(x=170, y=205)
button8.place(x= 250, y =205)
button9.place(x= 10, y =290)
button10.place(x=90, y=290)
button11.place(x= 170, y =290)
button12.place(x= 250, y =290)

inputFrame.pack()
window.mainloop()